# scripts/export_result_smartstore_to_excel.py
import json
import argparse
from pathlib import Path
from typing import Any, Dict, List, Tuple

from openpyxl import load_workbook


STANDARD_HEADERS = [
    "원스톱 상품번호",
    "원스톱 상품명",
    "카테고리",
    "상품명",
    "판매가",
    "옵션값",
    "옵션추가금액",
    "키워드",
    "대표이미지",
    "추가이미지1",
    "추가이미지2",
    "추가이미지3",
    "추가이미지4",
    "추가이미지5",
    "추가이미지6",
    "추가이미지7",
    "추가이미지8",
    "추가이미지9",
    "추가이미지10",
    "상세설명(HTML태그)",
]


def safe_get(obj: Any, *keys: str, default=None):
    cur = obj
    for key in keys:
        if not isinstance(cur, dict):
            return default
        cur = cur.get(key)
        if cur is None:
            return default
    return cur


def to_abs_onestop_url(url: str) -> str:
    if not url:
        return ""
    if url.startswith("http://") or url.startswith("https://"):
        return url
    if url.startswith("/"):
        return f"https://onestopdome.com{url}"
    return f"https://onestopdome.com/{url}"


def flatten_options(item: Dict[str, Any]) -> Tuple[str, str]:
    """
    표준양식:
    - 옵션값: 콤마(,) 구분
    - 옵션추가금액: 콤마(,) 구분

    현재 JSON 구조 예시:
    onestop.options = [
      {
        "optionName": "옵션",
        "values": [
          {"name": "혼합", "price": 11340, "diff": 4040},
          {"name": "실버", "price": 7300, "diff": 0}
        ]
      }
    ]
    """
    options = safe_get(item, "onestop", "options", default=[]) or []

    option_names: List[str] = []
    option_diffs: List[str] = []

    for group in options:
        values = group.get("values", []) or []
        for value in values:
            name = str(value.get("name", "")).strip()
            diff = value.get("diff", 0)

            if name:
                option_names.append(name)
                option_diffs.append(str(diff if diff is not None else 0))

    return ",".join(option_names), ",".join(option_diffs)


def extract_keywords(item: Dict[str, Any]) -> str:
    smart_keywords = safe_get(item, "sello", "smartKeywords", default=[]) or []
    coupang_keywords = safe_get(item, "sello", "coupangKeywords", default=[]) or []

    merged: List[str] = []
    seen = set()

    for kw in smart_keywords + coupang_keywords:
        kw = str(kw).strip()
        if kw and kw not in seen:
            seen.add(kw)
            merged.append(kw)

    return ",".join(merged)


def extract_images(item: Dict[str, Any]) -> List[str]:
    """
    표준양식 기준:
    - 대표이미지 1개
    - 추가이미지 10개

    우선순위:
    1) aliexpress.selected[].image
    2) onestop.thumbnailUrl
    3) matched.image
    """
    ali_selected = safe_get(item, "aliexpress", "selected", default=[]) or []

    images: List[str] = []
    seen = set()

    for row in ali_selected:
        image = str(row.get("image", "")).strip()
        if image and image not in seen:
            seen.add(image)
            images.append(image)

    if not images:
        fallback1 = str(safe_get(item, "onestop", "thumbnailUrl", default="") or "").strip()
        fallback2 = str(safe_get(item, "matched", "image", default="") or "").strip()

        for image in [fallback1, fallback2]:
            if image and image not in seen:
                seen.add(image)
                images.append(image)

    return images[:11]


def build_standard_row(item: Dict[str, Any]) -> Dict[str, Any]:
    option_values, option_add_prices = flatten_options(item)
    images = extract_images(item)

    category = (
            safe_get(item, "matched", "category", default="")
            or safe_get(item, "sello", "smartCandidate", "category", default="")
            or safe_get(item, "sello", "coupangCandidate", "category", default="")
            or safe_get(item, "onestop", "category", default="")
            or ""
    )

    product_name = (
            safe_get(item, "matched", "title", default="")
            or safe_get(item, "sello", "smartCandidate", "title", default="")
            or safe_get(item, "sello", "coupangCandidate", "title", default="")
            or safe_get(item, "onestop", "title", default="")
            or ""
    )

    sale_price = (
            safe_get(item, "matched", "price", default=None)
            or safe_get(item, "sello", "smartCandidate", "price", default=None)
            or safe_get(item, "sello", "coupangCandidate", "price", default=None)
            or safe_get(item, "onestop", "price", default="")
            or ""
    )

    row = {
        "원스톱 상품번호": safe_get(item, "onestop", "no", default=""),
        "원스톱 상품명": safe_get(item, "onestop", "title", default=""),
        "카테고리": category,
        "상품명": product_name,
        "판매가": sale_price,
        "옵션값": option_values,
        "옵션추가금액": option_add_prices,
        "키워드": extract_keywords(item),
        "대표이미지": images[0] if len(images) > 0 else "",
        "추가이미지1": images[1] if len(images) > 1 else "",
        "추가이미지2": images[2] if len(images) > 2 else "",
        "추가이미지3": images[3] if len(images) > 3 else "",
        "추가이미지4": images[4] if len(images) > 4 else "",
        "추가이미지5": images[5] if len(images) > 5 else "",
        "추가이미지6": images[6] if len(images) > 6 else "",
        "추가이미지7": images[7] if len(images) > 7 else "",
        "추가이미지8": images[8] if len(images) > 8 else "",
        "추가이미지9": images[9] if len(images) > 9 else "",
        "추가이미지10": images[10] if len(images) > 10 else "",
        "상세설명(HTML태그)": safe_get(item, "detailHtml", default="")
                        or safe_get(item, "onestop", "detailHtml", default="")
                        or "",
    }

    return row


def build_extra_row(item: Dict[str, Any]) -> Dict[str, Any]:
    onestop_url = to_abs_onestop_url(safe_get(item, "onestop", "url", default=""))
    onestop_final_url = to_abs_onestop_url(safe_get(item, "onestop", "finalUrl", default=""))

    ali_selected = safe_get(item, "aliexpress", "selected", default=[]) or []

    extra = {
        "createdAt": safe_get(item, "createdAt", default=""),
        "searchKeyword": safe_get(item, "searchKeyword", default=""),
        "onestop_exists": safe_get(item, "onestop", "exists", default=""),
        "onestop_url": onestop_url,
        "onestop_finalUrl": onestop_final_url,
        "onestop_priceText": safe_get(item, "onestop", "priceText", default=""),
        "onestop_thumbnailUrl": safe_get(item, "onestop", "thumbnailUrl", default=""),
        "onestop_category_raw": safe_get(item, "onestop", "category", default=""),
        "sello_rawCount": safe_get(item, "sello", "rawCount", default=""),
        "smartKeywords_raw": json.dumps(
            safe_get(item, "sello", "smartKeywords", default=[]), ensure_ascii=False
        ),
        "coupangKeywords_raw": json.dumps(
            safe_get(item, "sello", "coupangKeywords", default=[]), ensure_ascii=False
        ),
        "smartCandidate_title": safe_get(item, "sello", "smartCandidate", "title", default=""),
        "smartCandidate_link": safe_get(item, "sello", "smartCandidate", "link", default=""),
        "smartCandidate_mallName": safe_get(item, "sello", "smartCandidate", "mallName", default=""),
        "smartCandidate_price": safe_get(item, "sello", "smartCandidate", "price", default=""),
        "smartCandidate_category": safe_get(item, "sello", "smartCandidate", "category", default=""),
        "coupangCandidate_title": safe_get(item, "sello", "coupangCandidate", "title", default=""),
        "coupangCandidate_link": safe_get(item, "sello", "coupangCandidate", "link", default=""),
        "coupangCandidate_mallName": safe_get(item, "sello", "coupangCandidate", "mallName", default=""),
        "coupangCandidate_price": safe_get(item, "sello", "coupangCandidate", "price", default=""),
        "coupangCandidate_category": safe_get(item, "sello", "coupangCandidate", "category", default=""),
        "matched_title": safe_get(item, "matched", "title", default=""),
        "matched_link": safe_get(item, "matched", "link", default=""),
        "matched_mallName": safe_get(item, "matched", "mallName", default=""),
        "matched_price": safe_get(item, "matched", "price", default=""),
        "matched_category": safe_get(item, "matched", "category", default=""),
        "matched_productId": safe_get(item, "matched", "productId", default=""),
        "matched_productType": safe_get(item, "matched", "productType", default=""),
        "aliexpress_searched": safe_get(item, "aliexpress", "searched", default=""),
        "aliexpress_queryImage": safe_get(item, "aliexpress", "queryImage", default=""),
        "aliexpress_selected_count": len(ali_selected),
        "raw_json": json.dumps(item, ensure_ascii=False),
    }

    for idx, ali in enumerate(ali_selected[:10], start=1):
        extra[f"ali_{idx}_title"] = ali.get("title", "")
        extra[f"ali_{idx}_image"] = ali.get("image", "")
        extra[f"ali_{idx}_url"] = ali.get("url", "")
        extra[f"ali_{idx}_id"] = ali.get("id", "")

    return extra


def collect_all_extra_headers(items: List[Dict[str, Any]]) -> List[str]:
    headers: List[str] = []
    seen = set()

    for item in items:
        extra = build_extra_row(item)
        for key in extra.keys():
            if key not in seen:
                seen.add(key)
                headers.append(key)

    return headers


def copy_template_style(ws, source_row: int, target_row: int, max_col: int):
    """
    2행 스타일/높이를 데이터 행들에 복사
    """
    from copy import copy

    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height

    for col in range(1, max_col + 1):
        src = ws.cell(source_row, col)
        dst = ws.cell(target_row, col)

        if src.has_style:
            dst._style = copy(src._style)

        if src.number_format:
            dst.number_format = src.number_format

        if src.font:
            dst.font = copy(src.font)

        if src.fill:
            dst.fill = copy(src.fill)

        if src.border:
            dst.border = copy(src.border)

        if src.alignment:
            dst.alignment = copy(src.alignment)

        if src.protection:
            dst.protection = copy(src.protection)


def autofit_widths(ws, max_row: int, max_col: int):
    for col in range(1, max_col + 1):
        col_letter = ws.cell(1, col).column_letter
        max_len = 0

        for row in range(1, max_row + 1):
            value = ws.cell(row, col).value
            if value is None:
                continue

            text = str(value)
            for line in text.splitlines():
                if len(line) > max_len:
                    max_len = len(line)

        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 45)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--json", required=True, help="result_smartstore.json 경로")
    parser.add_argument("--template", required=True, help="표준양식 xlsx 경로")
    parser.add_argument("--output", required=True, help="출력 xlsx 경로")
    args = parser.parse_args()

    json_path = Path(args.json)
    template_path = Path(args.template)
    output_path = Path(args.output)

    with json_path.open("r", encoding="utf-8") as f:
        items = json.load(f)

    if not isinstance(items, list):
        raise ValueError("JSON 최상위 구조는 list 여야 합니다.")

    wb = load_workbook(template_path)
    ws = wb.active

    # A:T 표준 헤더 확인 / 필요시 덮어쓰기
    for idx, header in enumerate(STANDARD_HEADERS, start=1):
        ws.cell(row=1, column=idx, value=header)

    extra_headers = collect_all_extra_headers(items)

    # U열 이후 추가 헤더 작성
    extra_start_col = len(STANDARD_HEADERS) + 1
    for offset, header in enumerate(extra_headers, start=0):
        ws.cell(row=1, column=extra_start_col + offset, value=header)

    # 기존 설명행(2행)은 남기지 않고, 데이터를 2행부터 다시 씀
    # 필요하면 아래 주석을 해제해서 2행 샘플설명 유지 가능
    # pass

    for i, item in enumerate(items, start=2):
        standard_row = build_standard_row(item)
        extra_row = build_extra_row(item)

        copy_template_style(ws, source_row=2, target_row=i, max_col=len(STANDARD_HEADERS))

        # 표준 컬럼 입력
        for col_idx, header in enumerate(STANDARD_HEADERS, start=1):
            ws.cell(row=i, column=col_idx, value=standard_row.get(header, ""))

        # 추가 컬럼 입력
        for offset, header in enumerate(extra_headers, start=0):
            ws.cell(row=i, column=extra_start_col + offset, value=extra_row.get(header, ""))

    # 샘플 설명 행이 남아있고 데이터가 1건 이상이면 필요에 따라 삭제 가능
    # 템플릿 2행을 실제 첫 데이터로 덮어썼으므로 별도 삭제는 필요 없음

    autofit_widths(ws, max_row=ws.max_row, max_col=ws.max_column)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    print(f"완료: {output_path}")


if __name__ == "__main__":
    main()